"""XLSX importer for spectrum data.

Supports two formats:
1. Two-column format: wavelength + value (with optional header)
2. Multi-column format: wavelength + multiple spectra (header row required)
"""

from __future__ import annotations

from pathlib import Path

import numpy as np

from colorlab_pro.dto.spectrum import Spectrum


def _detect_channel_from_name(name: str) -> str | None:
    """Auto-detect channel (R/G/B) from spectrum name.

    Returns 'R', 'G', 'B', or None.

    Detection priority:
    1. Name ends with -R/-G/-B (e.g. BCP4011-R, LGC-R)
    2. Name starts with R/G/B followed by digit or dash (e.g. R110B-2, G130B-6)
    3. Name contains RED/GREEN/BLUE keyword
    """
    upper = name.upper().strip()

    # 1. Check if name ends with -R/-G/-B (after a separator)
    if upper.endswith("-R") or upper.endswith("_R"):
        return "R"
    if upper.endswith("-G") or upper.endswith("_G"):
        return "G"
    if upper.endswith("-B") or upper.endswith("_B"):
        return "B"

    # 2. Check if name starts with R/G/B followed by digit or dash
    if len(upper) >= 2:
        if upper[0] == "R" and upper[1] in "0123456789-":
            return "R"
        if upper[0] == "G" and upper[1] in "0123456789-":
            return "G"
        if upper[0] == "B" and upper[1] in "0123456789-":
            return "B"

    # 3. Simple R/G/B as entire name
    if upper == "R":
        return "R"
    if upper == "G":
        return "G"
    if upper == "B":
        return "B"

    # 4. Keywords
    if "RED" in upper:
        return "R"
    if "GREEN" in upper or "GRN" in upper:
        return "G"
    if "BLUE" in upper or "BLU" in upper:
        return "B"

    # 5. Name ends with R/G/B (e.g. BD1-SC1000R)
    if upper.endswith("R") and not upper.endswith("GR"):
        return "R"
    if upper.endswith("G"):
        return "G"
    if upper.endswith("B"):
        return "B"

    return None


def import_xlsx(
    path: Path,
    *,
    sheet: str | int = 0,
    unit: str = "a.u.",
    column: int | None = None,
    default_category: str | None = None,
) -> Spectrum | list[Spectrum]:
    """Load Spectrum(s) from an Excel file.

    Args:
        path: Path to the .xlsx file.
        sheet: Sheet name or index. Default is the first sheet.
        unit: Unit string for the spectrum values.
        column: If specified, import only this column (0-indexed, excluding wavelength).
                If None, import all columns as separate spectra.
        default_category: Force this category (CF/QD/LED/White) on all imported spectra.
                         If None, category is auto-detected from channel.

    Returns:
        A single Spectrum if column is specified, or a list of Spectra for multi-column.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the data cannot be parsed.
    """
    if not path.exists():
        raise FileNotFoundError(f"XLSX file not found: {path}")

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    if isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[sheet]

    all_rows: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        if row is not None and len(row) >= 2:
            all_rows.append(row)

    wb.close()

    if len(all_rows) < 2:
        raise ValueError("XLSX must contain at least 2 rows of data.")

    # Detect format: check if first row contains non-numeric headers
    first_row = all_rows[0]
    has_header = False
    try:
        float(first_row[0])
        float(first_row[1])
    except (ValueError, TypeError):
        has_header = True

    # Auto-detect the wavelength column (may not be column 0 if there are
    # empty leading columns, e.g. CF.xlsx with None in column A).
    # Strategy: for each column, check if its numeric values form a
    # monotonically increasing sequence (typical of wavelength data).
    # Among matching columns, pick the one with the most numeric values.
    max_cols = max(len(row) for row in all_rows)
    best_col = 0
    best_count = 0
    for col_idx in range(max_cols):
        numeric_vals: list[float] = []
        for row in all_rows:
            try:
                numeric_vals.append(float(row[col_idx]))
            except (ValueError, TypeError):
                numeric_vals.append(float("nan"))
        valid = [v for v in numeric_vals if v == v]  # filter NaN
        if len(valid) >= 2 and all(
            valid[i] < valid[i + 1] for i in range(len(valid) - 1)
        ):
            if len(valid) > best_count:
                best_count = len(valid)
                best_col = col_idx
    wl_col = best_col

    # Find where numeric data starts (skip header and any non-numeric rows like thickness)
    start_idx = 0
    for i, row in enumerate(all_rows):
        try:
            float(row[wl_col])
            float(row[wl_col + 1]) if wl_col + 1 < len(row) else None
            start_idx = i
            break
        except (ValueError, TypeError):
            continue

    # Extract wavelength column and valid data rows
    wavelengths: list[float] = []
    valid_rows: list[tuple] = []
    for row in all_rows[start_idx:]:
        try:
            wl = float(row[wl_col])
            wavelengths.append(wl)
            valid_rows.append(row)
        except (ValueError, TypeError):
            continue

    if len(wavelengths) < 2:
        raise ValueError("XLSX must contain at least 2 numeric wavelength rows.")

    wavelength_array = np.array(wavelengths)

    # Extract header names if present — scan all rows before the data to
    # find the actual header row (the one with the most string values in
    # the value columns, as opposed to thickness rows with numbers).
    headers: list[str] = []
    if has_header and start_idx > 0:
        best_header_idx = 0
        best_string_count = -1
        for hi in range(start_idx):
            row = all_rows[hi]
            str_count = sum(
                1
                for ci in range(wl_col + 1, len(row))
                if row[ci] is not None and not isinstance(row[ci], (int, float))
            )
            if str_count > best_string_count:
                best_string_count = str_count
                best_header_idx = hi
        header_row = all_rows[best_header_idx]
        headers = [
            str(h) if h is not None else f"Column_{i}"
            for i, h in enumerate(header_row[wl_col + 1:], 1)
        ]

    # Determine number of value columns from ALL data rows (not just first).
    # Some rows may have trailing None cells that reduce len(row); scanning
    # all rows and taking the max ensures we don't miss columns.
    num_value_cols = 0
    for row in valid_rows:
        row_len = len(row) - wl_col - 1  # Exclude wavelength column and leading empty cols
        if row_len > num_value_cols:
            num_value_cols = row_len

    # Extract thickness from a row between header and data (if present).
    # The thickness row has a non-numeric label in the wl_col but numeric
    # values in the value columns (e.g. "THX/um", 2.2, 2.2, 2.2).
    thickness_values: list[float | None] = []
    if has_header and start_idx > 1:
        thickness_row = None
        for ti in range(1, start_idx):
            row = all_rows[ti]
            try:
                float(row[wl_col])
                continue  # numeric in wl_col — data row or other, skip
            except (ValueError, TypeError):
                pass
            # Check if this row has numeric values in the value columns
            num_count = sum(
                1
                for ci in range(wl_col + 1, len(row))
                if isinstance(row[ci], (int, float))
            )
            if num_count >= 1:
                thickness_row = row
                break
        if thickness_row is not None:
            for col_idx in range(num_value_cols):
                try:
                    val = thickness_row[wl_col + col_idx + 1] if wl_col + col_idx + 1 < len(thickness_row) else None
                    thickness_values.append(float(val) if val is not None else None)
                except (ValueError, TypeError):
                    thickness_values.append(None)

    # Extract value columns
    spectra: list[Spectrum] = []
    for col_idx in range(num_value_cols):
        if column is not None and col_idx != column:
            continue

        values: list[float] = []
        for row in valid_rows:
            try:
                val_idx = wl_col + col_idx + 1
                if val_idx < len(row):
                    val = row[val_idx]
                    values.append(float(val) if val is not None else 0.0)
                else:
                    values.append(0.0)
            except (ValueError, TypeError):
                values.append(0.0)

        name = headers[col_idx] if col_idx < len(headers) else f"Spectrum_{col_idx + 1}"

        # Auto-detect channel from name
        channel = _detect_channel_from_name(name)

        # Get thickness for this column
        thickness = thickness_values[col_idx] if col_idx < len(thickness_values) else None

        meta: dict[str, object] = {
            "source_file": str(path.name),
            "name": name,
            "column_index": col_idx,
            "channel": channel,
        }
        if default_category is not None:
            meta["category"] = default_category
        if thickness is not None:
            meta["thickness_um"] = thickness
        else:
            meta["thickness_um"] = None
            meta["thickness_missing"] = True

        spectra.append(
            Spectrum(
                wavelengths=wavelength_array.copy(),
                values=np.array(values),
                unit=unit,
                meta=meta,
            )
        )

    if column is not None:
        return spectra[0] if spectra else None

    # For single-column files (wavelength + 1 value), return single Spectrum
    if len(spectra) == 1:
        return spectra[0]

    return spectra
