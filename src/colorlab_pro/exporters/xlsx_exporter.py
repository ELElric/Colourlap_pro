"""XLSX import/export for Spectrum DTOs."""

from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook

from colorlab_pro.dto.spectrum import Spectrum

__all__ = ["export_spectrum", "import_spectrum"]


def export_spectrum(spectrum: Spectrum, path: Path) -> None:
    """Write a spectrum to an XLSX file with wavelength and value columns."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        sheet = workbook.create_sheet("Spectrum")
    sheet.title = "Spectrum"
    sheet.append(["wavelength_nm", "value", "unit"])
    for wl, val in zip(spectrum.wavelengths, spectrum.values, strict=True):
        sheet.append([float(wl), float(val), spectrum.unit])
    workbook.save(path)


def import_spectrum(path: Path, *, unit: str = "a.u.") -> Spectrum:
    """Read a spectrum from an XLSX file."""
    workbook = load_workbook(path)
    sheet = workbook.active
    if sheet is None:
        raise ValueError("Workbook has no active sheet")

    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("At least two rows are required (header + data)")

    header = rows[0]
    has_unit = "unit" in [str(h).strip().lower() for h in header if h is not None]
    unit_index = next(
        (i for i, h in enumerate(header) if str(h).strip().lower() == "unit"),
        None,
    )

    wavelengths: list[float] = []
    values: list[float] = []
    file_unit = unit

    for row in rows[1:]:
        if row is None or not row or row[0] is None:
            continue
        try:
            wavelengths.append(float(row[0]))
            values.append(float(row[1]))
        except (ValueError, TypeError, IndexError) as exc:
            raise ValueError(f"Invalid numeric row: {row}") from exc
        if has_unit and unit_index is not None and len(row) > unit_index:
            file_unit = str(row[unit_index] or "").strip() or unit

    if len(wavelengths) < 2:
        raise ValueError("At least two data points are required")

    return Spectrum(
        wavelengths=np.array(wavelengths, dtype=np.float64),
        values=np.array(values, dtype=np.float64),
        unit=file_unit,
    )
