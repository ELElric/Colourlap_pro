"""Unit tests for XLSX spectrum importer."""

from __future__ import annotations

import numpy as np
import pytest
from openpyxl import Workbook

from colorlab_pro.dto.spectrum import Spectrum
from colorlab_pro.importers.xlsx_importer import import_xlsx


def _write_xlsx(path, rows):
    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        sheet = workbook.create_sheet("Spectrum")
    sheet.title = "Spectrum"
    for row in rows:
        sheet.append(row)
    workbook.save(path)


@pytest.fixture
def sample_xlsx(tmp_path):
    path = tmp_path / "sample.xlsx"
    _write_xlsx(
        path,
        [
            ["wavelength_nm", "value"],
            [400.0, 0.1],
            [500.0, 0.5],
            [600.0, 0.9],
        ],
    )
    return path


def test_import_xlsx_with_header(sample_xlsx):
    spectrum = import_xlsx(sample_xlsx)
    assert isinstance(spectrum, Spectrum)
    np.testing.assert_array_equal(spectrum.wavelengths, np.array([400.0, 500.0, 600.0]))
    np.testing.assert_array_equal(spectrum.values, np.array([0.1, 0.5, 0.9]))
    assert spectrum.unit == "a.u."
    assert spectrum.meta.get("source_file") == "sample.xlsx"


def test_import_xlsx_without_header(tmp_path):
    path = tmp_path / "no_header.xlsx"
    _write_xlsx(
        path,
        [
            [400.0, 0.1],
            [500.0, 0.5],
            [600.0, 0.9],
        ],
    )
    spectrum = import_xlsx(path)
    np.testing.assert_array_equal(spectrum.wavelengths, np.array([400.0, 500.0, 600.0]))
    np.testing.assert_array_equal(spectrum.values, np.array([0.1, 0.5, 0.9]))


def test_import_xlsx_by_sheet_name(tmp_path):
    path = tmp_path / "sheet_name.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        sheet = workbook.create_sheet("Spectrum")
    sheet.title = "Spectrum"
    sheet.append(["wavelength_nm", "value"])
    sheet.append([400.0, 0.1])
    sheet.append([500.0, 0.5])
    sheet.append([600.0, 0.9])
    workbook.save(path)

    spectrum = import_xlsx(path, sheet="Spectrum")
    np.testing.assert_array_equal(spectrum.wavelengths, np.array([400.0, 500.0, 600.0]))


def test_import_xlsx_file_not_found(tmp_path):
    path = tmp_path / "missing.xlsx"
    with pytest.raises(FileNotFoundError):
        import_xlsx(path)


def test_import_xlsx_insufficient_rows(tmp_path):
    path = tmp_path / "too_few.xlsx"
    _write_xlsx(path, [["wavelength_nm", "value"], [400.0, 0.1]])
    with pytest.raises(ValueError, match="at least 2 numeric"):
        import_xlsx(path)


def test_import_xlsx_only_header(tmp_path):
    path = tmp_path / "header_only.xlsx"
    _write_xlsx(path, [["wavelength_nm", "value"]])
    with pytest.raises(ValueError, match="at least 2"):
        import_xlsx(path)


def test_import_xlsx_skips_non_numeric_rows(tmp_path):
    path = tmp_path / "with_gaps.xlsx"
    _write_xlsx(
        path,
        [
            ["wavelength_nm", "value"],
            [400.0, 0.1],
            ["bad", "row"],
            [500.0, 0.5],
            [600.0, 0.9],
        ],
    )
    spectrum = import_xlsx(path)
    np.testing.assert_array_equal(spectrum.wavelengths, np.array([400.0, 500.0, 600.0]))
    np.testing.assert_array_equal(spectrum.values, np.array([0.1, 0.5, 0.9]))


def test_import_xlsx_custom_unit(tmp_path):
    path = tmp_path / "custom_unit.xlsx"
    _write_xlsx(
        path,
        [
            ["wavelength_nm", "value"],
            [400.0, 0.1],
            [500.0, 0.5],
            [600.0, 0.9],
        ],
    )
    spectrum = import_xlsx(path, unit="mW/nm")
    assert spectrum.unit == "mW/nm"
