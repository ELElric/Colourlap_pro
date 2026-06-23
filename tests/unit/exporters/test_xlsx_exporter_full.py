"""Additional unit tests for XLSX exporter/importer."""

from __future__ import annotations

from unittest.mock import MagicMock, patch

import numpy as np
import pytest
from openpyxl import Workbook, load_workbook

from colorlab_pro.dto.spectrum import Spectrum
from colorlab_pro.exporters.xlsx_exporter import export_spectrum, import_spectrum


@pytest.fixture
def sample_spectrum() -> Spectrum:
    return Spectrum(
        wavelengths=np.array([400.0, 500.0, 600.0]),
        values=np.array([0.1, 0.5, 0.9]),
        unit="mW/nm",
    )


def test_export_spectrum_creates_file(tmp_path, sample_spectrum):
    path = tmp_path / "nested" / "spectrum.xlsx"
    export_spectrum(sample_spectrum, path)
    assert path.exists()
    assert path.parent.exists()


def test_export_spectrum_creates_sheet_when_active_none(tmp_path, sample_spectrum):
    path = tmp_path / "no_active.xlsx"
    mock_workbook = MagicMock()
    mock_workbook.active = None
    mock_sheet = MagicMock()
    mock_workbook.create_sheet.return_value = mock_sheet
    with patch("colorlab_pro.exporters.xlsx_exporter.Workbook", return_value=mock_workbook):
        export_spectrum(sample_spectrum, path)
    mock_workbook.create_sheet.assert_called_once_with("Spectrum")
    assert mock_sheet.title == "Spectrum"


def test_import_empty_workbook(tmp_path):
    path = tmp_path / "empty.xlsx"
    workbook = Workbook()
    workbook.active.append(["wavelength_nm", "value"])
    workbook.save(path)
    with pytest.raises(ValueError, match="At least two rows are required"):
        import_spectrum(path)


def test_import_invalid_numeric_row(tmp_path):
    path = tmp_path / "invalid.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["wavelength_nm", "value"])
    sheet.append([400.0, 0.1])
    sheet.append([500.0, "bad"])
    workbook.save(path)
    with pytest.raises(ValueError, match="Invalid numeric row"):
        import_spectrum(path)


def test_import_missing_unit_column(tmp_path):
    path = tmp_path / "no_unit.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["wavelength_nm", "value"])
    sheet.append([400.0, 0.1])
    sheet.append([500.0, 0.5])
    sheet.append([600.0, 0.9])
    workbook.save(path)
    spectrum = import_spectrum(path)
    assert spectrum.unit == "a.u."


def test_import_with_unit_column(tmp_path):
    path = tmp_path / "with_unit.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["wavelength_nm", "value", "unit"])
    sheet.append([400.0, 0.1, "mW/nm"])
    sheet.append([500.0, 0.5, "mW/nm"])
    sheet.append([600.0, 0.9, "mW/nm"])
    workbook.save(path)
    spectrum = import_spectrum(path)
    assert spectrum.unit == "mW/nm"


def test_import_with_unit_column_blank_fallback(tmp_path):
    path = tmp_path / "blank_unit.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["wavelength_nm", "value", "unit"])
    sheet.append([400.0, 0.1, ""])
    sheet.append([500.0, 0.5, ""])
    sheet.append([600.0, 0.9, ""])
    workbook.save(path)
    spectrum = import_spectrum(path, unit="fallback")
    assert spectrum.unit == "fallback"


def test_import_too_few_points(tmp_path):
    path = tmp_path / "too_few.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["wavelength_nm", "value"])
    sheet.append([400.0, 0.1])
    workbook.save(path)
    with pytest.raises(ValueError, match="At least two data points are required"):
        import_spectrum(path)


def test_import_no_active_sheet(tmp_path):
    path = tmp_path / "no_active.xlsx"
    mock_workbook = Workbook()
    mock_workbook.remove(mock_workbook.active)
    with patch("colorlab_pro.exporters.xlsx_exporter.load_workbook", return_value=mock_workbook):
        with pytest.raises(ValueError, match="no active sheet"):
            import_spectrum(path)


def test_import_ignores_blank_rows(tmp_path):
    path = tmp_path / "blanks.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["wavelength_nm", "value"])
    sheet.append([400.0, 0.1])
    sheet.append([])
    sheet.append([500.0, 0.5])
    sheet.append([600.0, 0.9])
    workbook.save(path)
    spectrum = import_spectrum(path)
    np.testing.assert_array_equal(spectrum.wavelengths, np.array([400.0, 500.0, 600.0]))


def test_export_import_preserves_unit_default(tmp_path):
    spectrum = Spectrum(
        wavelengths=np.array([400.0, 500.0]),
        values=np.array([0.1, 0.5]),
    )
    path = tmp_path / "default_unit.xlsx"
    export_spectrum(spectrum, path)
    loaded = import_spectrum(path)
    assert loaded.unit == "a.u."


def test_exported_xlsx_format(tmp_path, sample_spectrum):
    path = tmp_path / "format.xlsx"
    export_spectrum(sample_spectrum, path)
    loaded = load_workbook(path)
    sheet = loaded.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("wavelength_nm", "value", "unit")
    assert rows[1] == (400.0, 0.1, "mW/nm")
